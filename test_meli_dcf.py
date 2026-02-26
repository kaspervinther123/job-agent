#!/usr/bin/env python3
"""
Test suite for MELI_Reverse_DCF_Model.xlsx
Run: python3 test_meli_dcf.py
"""
import openpyxl, sys

PASS="\033[92m✓\033[0m"; FAIL="\033[91m✗\033[0m"
results=[]
def check(n,c,d=""):
    ok=bool(c); results.append((ok,n,d))
    print(f"  {PASS if ok else FAIL}  {n}"+(f"  ({d})" if d else ""))

wb=openpyxl.load_workbook("MELI_Reverse_DCF_Model.xlsx")

print("\n=== Sheets ===")
for s in ["Dashboard","Reverse DCF","WACC","Historicals","Sensitivity","Instructions"]:
    check(f"'{s}' present", s in wb.sheetnames)

print("\n=== Reverse DCF inputs ===")
ws=wb["Reverse DCF"]
check("C8 BaseFCF=3800",  ws["C8"].value==3800)
check("C9 Shares=51",     ws["C9"].value==51.0)
check("C10 NetDebt=2000", ws["C10"].value==2000)
check("C14 TermG=4%",     ws["C14"].value==0.04)
check("C15 ProjYrs=10",   ws["C15"].value==10)
check("C16 TargetG=25%",  ws["C16"].value==0.25)
check("C7 STOCKHISTORY",  "STOCKHISTORY" in str(ws["C7"].value or ""))
check("C13 links WACC",   "WACC" in str(ws["C13"].value or ""))

print("\n=== Lookup table ===")
check("H8=0%",  ws["H8"].value==0.0)
check("H68=60%",ws["H68"].value==0.6)
check("61 rows",sum(1 for r in range(8,69) if ws.cell(r,8).value is not None)==61)
i8=str(ws["I8"].value or "")
for ref in ["$C$8","$C$13","$C$14","$C$15","SUMPRODUCT"]:
    check(f"I8 has {ref}", ref in i8)
check("J8 has $C$7", "$C$7" in str(ws["J8"].value or ""))

print("\n=== Implied growth formula F7 ===")
f7=str(ws["F7"].value or "")
for kw in ["INDEX","MATCH","$H$8","$H$68","$I$8","$I$68","$C$7","ABS"]:
    check(f"F7 has {kw}", kw in f7)

print("\n=== Projection rows 30-39 ===")
for yr in range(1,11):
    d=str(ws.cell(29+yr,4).value or "")
    check(f"Year {yr} FCF formula", f"$C$8*(1+$F$7)^{yr}" in d)

print("\n=== Equity bridge rows 41-48 ===")
for row,frag in [(41,"SUM(G30:G39)"),(42,"IFERROR"),(43,"(1+$C$13)"),
                 (44,"=C41+C43"),(45,"$C$10"),(46,"=C44-C45"),(47,"$C$9"),(48,"C46")]:
    check(f"C{row} bridge formula", frag in str(ws.cell(row,3).value or ""))

print("\n=== WACC sheet ===")
ww=wb["WACC"]
for cell,expected in [("C7",0.045),("C8",1.35),("C9",0.055),("C10",0.015),
                       ("C15",0.065),("C16",0.30),("C21",97000),("C22",7000)]:
    check(f"{cell}={expected}", ww[cell].value==expected)
wf=str(ww["C28"].value or "")
for part in ["C24","C11","C25","C17"]:
    check(f"WACC formula has {part}", part in wf)
check("Ke: C7+C8*(C9+C10)", "C7+C8*(C9+C10)" in str(ww["C11"].value or ""))
check("Kd: C15*(1-C16)",    "C15*(1-C16)"    in str(ww["C17"].value or ""))

print("\n=== Historicals ===")
wh=wb["Historicals"]
for j,yr in enumerate([2020,2021,2022,2023,2024]):
    check(f"Year col {3+j}={yr}", str(wh.cell(5,3+j).value)==str(yr))
for r in range(6,30):
    if str(wh.cell(r,2).value or "").strip().lower()=="revenue":
        check("FY2023 Rev=$14,474M", wh.cell(r,6).value==14474)
        check("FY2024 Rev=$20,777M", wh.cell(r,7).value==20777)
        break

print("\n=== Sensitivity ===")
ws_s=wb["Sensitivity"]
for row,wstr in [(8,"8%"),(9,"9%"),(10,"10%"),(16,"16%")]:
    check(f"Row {row} WACC='{wstr}'", ws_s.cell(row,2).value==wstr)
c8s=str(ws_s.cell(8,3).value or "")
check("Sens C8 SUMPRODUCT", "SUMPRODUCT" in c8s)
check("Sens C8 BaseFCF ref", "$C$8" in c8s)
check("Sens C8 WACC row $B8","$B8" in c8s)

print("\n=== DCF math sanity ===")
def dcf(g,wacc=0.115,tg=0.04,fcf=3800,nd=2000,sh=51.0,n=10):
    pv=sum(fcf*(1+g)**t/(1+wacc)**t for t in range(1,n+1))
    tv=fcf*(1+g)**n*(1+tg)/(wacc-tg)/(1+wacc)**n
    return (pv+tv-nd)/sh
p10,p20,p25,p30=dcf(0.10),dcf(0.20),dcf(0.25),dcf(0.30)
check("Monotone 10<20<25<30%", p10<p20<p25<p30)
lo,hi=0.0,0.60
for _ in range(80): mid=(lo+hi)/2; (hi:=mid) if dcf(mid)>1950 else (lo:=mid)
imp=(lo+hi)/2
check(f"Implied growth {imp:.2%} for $1,950 (0.1<g<0.6)", 0.10<imp<0.60)
check(f"DCF(implied)=$1,950 within $0.01", abs(dcf(imp)-1950)<0.01)

passed=sum(1 for ok,*_ in results if ok); total=len(results)
print(f"\n{'═'*55}")
print(f"  {passed}/{total} passed  {'ALL PASS ✓' if passed==total else 'FAIL'}")
print(f"{'═'*55}")
if passed<total:
    print("Failed:"); [print(f"  ✗  {n}  ({d})") for ok,n,d in results if not ok]
sys.exit(0 if passed==total else 1)
