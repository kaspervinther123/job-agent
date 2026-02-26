"""Excel-based visualization for podcast sentiment analysis."""

from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, BarChart3D
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

from .fetcher import Episode
from .sentiment import SentimentResult

EMOTION_COLORS = {
    "joy": "FFD700",       # Gold
    "trust": "228B22",     # Forest green
    "anticipation": "FF8C00",  # Dark orange
    "surprise": "9370DB",  # Medium purple
    "fear": "4B0082",      # Indigo
    "sadness": "4682B4",   # Steel blue
    "anger": "DC143C",     # Crimson
    "disgust": "8B4513",   # Saddle brown
}

HEADER_FILL = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
POS_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
NEG_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
NEU_FILL = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _style_header(ws, row: int, num_cols: int):
    """Apply header styling to a row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def _auto_width(ws, min_width: int = 10, max_width: int = 50):
    """Auto-adjust column widths."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, max_len + 2))


@dataclass
class AnalyzedEpisode:
    episode: Episode
    sentiment: SentimentResult


def create_report(
    analyzed: list[AnalyzedEpisode],
    output_path: str,
    podcast_name: str = "Millionærklubben",
) -> str:
    """Create an Excel report with sentiment analysis and charts."""
    wb = Workbook()

    _create_summary_sheet(wb, analyzed, podcast_name)
    _create_episodes_sheet(wb, analyzed)
    _create_monthly_sheet(wb, analyzed)
    _create_yearly_sheet(wb, analyzed)
    _create_emotions_sheet(wb, analyzed)

    wb.save(output_path)
    return output_path


def _create_summary_sheet(wb: Workbook, analyzed: list[AnalyzedEpisode], name: str):
    """Dashboard summary sheet."""
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = "2F4F4F"

    # Title
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"Podcast Sentiment Tracker — {name}"
    title_cell.font = Font(bold=True, size=16, color="2F4F4F")
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:H2")
    ws["A2"].value = f"Analysis of {len(analyzed)} episodes"
    ws["A2"].font = Font(size=11, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Key metrics
    row = 4
    metrics_header = ["Metric", "Value"]
    for c, h in enumerate(metrics_header, 1):
        ws.cell(row=row, column=c, value=h)
    _style_header(ws, row, len(metrics_header))

    if not analyzed:
        ws.cell(row=row + 1, column=1, value="No data available")
        return

    all_scores = [a.sentiment.polarity_score for a in analyzed]
    pos_count = sum(1 for a in analyzed if a.sentiment.label == "positive")
    neg_count = sum(1 for a in analyzed if a.sentiment.label == "negative")
    neu_count = sum(1 for a in analyzed if a.sentiment.label == "neutral")

    date_range = f"{analyzed[0].episode.pub_date.strftime('%Y-%m-%d')} to {analyzed[-1].episode.pub_date.strftime('%Y-%m-%d')}"

    metrics = [
        ("Total Episodes", len(analyzed)),
        ("Date Range", date_range),
        ("Average Sentiment", round(sum(all_scores) / len(all_scores), 3)),
        ("Most Positive Score", round(max(all_scores), 3)),
        ("Most Negative Score", round(min(all_scores), 3)),
        ("Positive Episodes", f"{pos_count} ({pos_count/len(analyzed)*100:.1f}%)"),
        ("Neutral Episodes", f"{neu_count} ({neu_count/len(analyzed)*100:.1f}%)"),
        ("Negative Episodes", f"{neg_count} ({neg_count/len(analyzed)*100:.1f}%)"),
    ]

    # Dominant emotion
    total_emotions = defaultdict(float)
    for a in analyzed:
        for emo, score in a.sentiment.emotions.items():
            total_emotions[emo] += score
    if total_emotions:
        dominant = max(total_emotions, key=total_emotions.get)
        metrics.append(("Dominant Emotion", dominant.capitalize()))

    for i, (metric, value) in enumerate(metrics):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=metric).font = Font(bold=True)
        ws.cell(row=r, column=2, value=value)
        for c in range(1, 3):
            ws.cell(row=r, column=c).border = THIN_BORDER

    # Emotion breakdown summary
    emo_row = row + len(metrics) + 3
    ws.cell(row=emo_row, column=1, value="Emotion Breakdown (Average Intensity)")
    ws.cell(row=emo_row, column=1).font = Font(bold=True, size=12, color="2F4F4F")

    emo_header_row = emo_row + 1
    emotions = list(EMOTION_COLORS.keys())
    for c, emo in enumerate(emotions, 1):
        ws.cell(row=emo_header_row, column=c, value=emo.capitalize())
    _style_header(ws, emo_header_row, len(emotions))

    for c, emo in enumerate(emotions, 1):
        avg = total_emotions[emo] / len(analyzed) if analyzed else 0
        cell = ws.cell(row=emo_header_row + 1, column=c, value=round(avg, 4))
        cell.number_format = "0.0000"
        cell.fill = PatternFill(
            start_color=EMOTION_COLORS[emo],
            end_color=EMOTION_COLORS[emo],
            fill_type="solid",
        )
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Emotion bar chart on the dashboard
    chart_row = emo_header_row + 4
    chart = BarChart()
    chart.type = "col"
    chart.title = "Average Emotion Intensity"
    chart.y_axis.title = "Intensity"
    chart.x_axis.title = "Emotion"
    chart.style = 10

    data_ref = Reference(ws, min_col=1, max_col=len(emotions),
                         min_row=emo_header_row + 1, max_row=emo_header_row + 1)
    cats_ref = Reference(ws, min_col=1, max_col=len(emotions),
                         min_row=emo_header_row, max_row=emo_header_row)
    chart.add_data(data_ref, from_rows=True)
    chart.set_categories(cats_ref)
    chart.shape = 4
    chart.width = 24
    chart.height = 14

    # Color each bar
    if chart.series:
        series = chart.series[0]
        for i, emo in enumerate(emotions):
            from openpyxl.chart.series import DataPoint
            from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = EMOTION_COLORS[emo]
            series.data_points.append(pt)

    chart.legend = None
    ws.add_chart(chart, f"A{chart_row}")

    _auto_width(ws)


def _create_episodes_sheet(wb: Workbook, analyzed: list[AnalyzedEpisode]):
    """Sheet with per-episode sentiment data."""
    ws = wb.create_sheet("Episodes")
    ws.sheet_properties.tabColor = "4682B4"

    emotions = list(EMOTION_COLORS.keys())
    headers = ["Date", "Title", "Polarity", "Label", "Raw Score",
               "Words", "Matched"] + [e.capitalize() for e in emotions]

    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    for i, a in enumerate(analyzed):
        row = i + 2
        ws.cell(row=row, column=1, value=a.episode.pub_date.strftime("%Y-%m-%d"))
        ws.cell(row=row, column=2, value=a.episode.title[:100])
        ws.cell(row=row, column=3, value=round(a.sentiment.polarity_score, 4))
        ws.cell(row=row, column=4, value=a.sentiment.label)
        ws.cell(row=row, column=5, value=a.sentiment.raw_score)
        ws.cell(row=row, column=6, value=a.sentiment.word_count)
        ws.cell(row=row, column=7, value=a.sentiment.matched_words)

        for j, emo in enumerate(emotions):
            ws.cell(row=row, column=8 + j,
                    value=round(a.sentiment.emotions.get(emo, 0), 4))

        # Color-code by sentiment
        fill = POS_FILL if a.sentiment.label == "positive" else (
            NEG_FILL if a.sentiment.label == "negative" else NEU_FILL)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = fill
            cell.border = THIN_BORDER

    # Polarity time series chart
    if len(analyzed) > 1:
        chart = LineChart()
        chart.title = "Sentiment Polarity Over Time"
        chart.y_axis.title = "Polarity (-1 to +1)"
        chart.x_axis.title = "Episode"
        chart.style = 10
        chart.width = 30
        chart.height = 15

        data = Reference(ws, min_col=3, min_row=1, max_row=len(analyzed) + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(analyzed) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        series = chart.series[0]
        series.graphicalProperties.line.solidFill = "2F4F4F"
        series.graphicalProperties.line.width = 15000  # in EMUs
        series.smooth = True

        ws.add_chart(chart, f"A{len(analyzed) + 4}")

    _auto_width(ws)
    ws.freeze_panes = "A2"


def _create_monthly_sheet(wb: Workbook, analyzed: list[AnalyzedEpisode]):
    """Monthly aggregated sentiment."""
    ws = wb.create_sheet("Monthly Trends")
    ws.sheet_properties.tabColor = "FF8C00"

    # Aggregate by month
    monthly: dict[str, list[AnalyzedEpisode]] = defaultdict(list)
    for a in analyzed:
        key = a.episode.pub_date.strftime("%Y-%m")
        monthly[key].append(a)

    emotions = list(EMOTION_COLORS.keys())
    headers = ["Month", "Episodes", "Avg Polarity", "Positive %", "Negative %"] + \
              [f"Avg {e.capitalize()}" for e in emotions]

    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    sorted_months = sorted(monthly.keys())
    for i, month in enumerate(sorted_months):
        row = i + 2
        eps = monthly[month]
        scores = [a.sentiment.polarity_score for a in eps]
        pos_pct = sum(1 for a in eps if a.sentiment.label == "positive") / len(eps) * 100
        neg_pct = sum(1 for a in eps if a.sentiment.label == "negative") / len(eps) * 100

        ws.cell(row=row, column=1, value=month)
        ws.cell(row=row, column=2, value=len(eps))
        ws.cell(row=row, column=3, value=round(sum(scores) / len(scores), 4))
        ws.cell(row=row, column=4, value=round(pos_pct, 1))
        ws.cell(row=row, column=5, value=round(neg_pct, 1))

        for j, emo in enumerate(emotions):
            avg_emo = sum(a.sentiment.emotions.get(emo, 0) for a in eps) / len(eps)
            ws.cell(row=row, column=6 + j, value=round(avg_emo, 4))

        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).border = THIN_BORDER

    # Monthly polarity trend chart
    if len(sorted_months) > 1:
        chart = LineChart()
        chart.title = "Monthly Average Sentiment"
        chart.y_axis.title = "Average Polarity"
        chart.x_axis.title = "Month"
        chart.style = 10
        chart.width = 30
        chart.height = 15

        data = Reference(ws, min_col=3, min_row=1, max_row=len(sorted_months) + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(sorted_months) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.series[0].graphicalProperties.line.solidFill = "FF8C00"
        chart.series[0].graphicalProperties.line.width = 20000
        chart.series[0].smooth = True

        ws.add_chart(chart, f"A{len(sorted_months) + 4}")

    # Positive/negative % stacked bar chart
    if len(sorted_months) > 1:
        bar = BarChart()
        bar.type = "col"
        bar.grouping = "stacked"
        bar.title = "Monthly Sentiment Distribution (%)"
        bar.y_axis.title = "Percentage"
        bar.style = 10
        bar.width = 30
        bar.height = 15

        pos_data = Reference(ws, min_col=4, min_row=1, max_row=len(sorted_months) + 1)
        neg_data = Reference(ws, min_col=5, min_row=1, max_row=len(sorted_months) + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(sorted_months) + 1)
        bar.add_data(pos_data, titles_from_data=True)
        bar.add_data(neg_data, titles_from_data=True)
        bar.set_categories(cats)
        bar.series[0].graphicalProperties.solidFill = "4CAF50"
        bar.series[1].graphicalProperties.solidFill = "F44336"

        ws.add_chart(bar, f"A{len(sorted_months) + 22}")

    # Monthly emotions stacked area chart
    if len(sorted_months) > 1:
        emo_chart = LineChart()  # Use line chart since area is simulated
        emo_chart.title = "Monthly Emotion Trends"
        emo_chart.y_axis.title = "Average Intensity"
        emo_chart.style = 10
        emo_chart.width = 30
        emo_chart.height = 15

        for j, emo in enumerate(emotions):
            col = 6 + j
            data = Reference(ws, min_col=col, min_row=1, max_row=len(sorted_months) + 1)
            emo_chart.add_data(data, titles_from_data=True)
            emo_chart.series[j].graphicalProperties.line.solidFill = EMOTION_COLORS[emo]
            emo_chart.series[j].graphicalProperties.line.width = 15000

        cats = Reference(ws, min_col=1, min_row=2, max_row=len(sorted_months) + 1)
        emo_chart.set_categories(cats)

        ws.add_chart(emo_chart, f"A{len(sorted_months) + 40}")

    _auto_width(ws)
    ws.freeze_panes = "A2"


def _create_yearly_sheet(wb: Workbook, analyzed: list[AnalyzedEpisode]):
    """Yearly aggregated sentiment."""
    ws = wb.create_sheet("Yearly Overview")
    ws.sheet_properties.tabColor = "228B22"

    yearly: dict[int, list[AnalyzedEpisode]] = defaultdict(list)
    for a in analyzed:
        yearly[a.episode.pub_date.year].append(a)

    emotions = list(EMOTION_COLORS.keys())
    headers = ["Year", "Episodes", "Avg Polarity", "Positive %",
               "Neutral %", "Negative %"] + [e.capitalize() for e in emotions]

    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    sorted_years = sorted(yearly.keys())
    for i, year in enumerate(sorted_years):
        row = i + 2
        eps = yearly[year]
        scores = [a.sentiment.polarity_score for a in eps]
        total = len(eps)

        ws.cell(row=row, column=1, value=year)
        ws.cell(row=row, column=2, value=total)
        ws.cell(row=row, column=3, value=round(sum(scores) / total, 4))
        ws.cell(row=row, column=4,
                value=round(sum(1 for a in eps if a.sentiment.label == "positive") / total * 100, 1))
        ws.cell(row=row, column=5,
                value=round(sum(1 for a in eps if a.sentiment.label == "neutral") / total * 100, 1))
        ws.cell(row=row, column=6,
                value=round(sum(1 for a in eps if a.sentiment.label == "negative") / total * 100, 1))

        for j, emo in enumerate(emotions):
            avg_emo = sum(a.sentiment.emotions.get(emo, 0) for a in eps) / total
            ws.cell(row=row, column=7 + j, value=round(avg_emo, 4))

        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).border = THIN_BORDER

    # Yearly polarity bar chart
    if len(sorted_years) > 1:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Yearly Average Sentiment"
        chart.y_axis.title = "Average Polarity"
        chart.style = 10
        chart.width = 24
        chart.height = 14

        data = Reference(ws, min_col=3, min_row=1, max_row=len(sorted_years) + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(sorted_years) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # Color bars based on positive/negative
        series = chart.series[0]
        for i, year in enumerate(sorted_years):
            from openpyxl.chart.series import DataPoint
            pt = DataPoint(idx=i)
            eps = yearly[year]
            avg = sum(a.sentiment.polarity_score for a in eps) / len(eps)
            pt.graphicalProperties.solidFill = "4CAF50" if avg >= 0 else "F44336"
            series.data_points.append(pt)

        chart.legend = None
        ws.add_chart(chart, f"A{len(sorted_years) + 4}")

    # Yearly emotion grouped bar chart
    if len(sorted_years) > 1:
        emo_chart = BarChart()
        emo_chart.type = "col"
        emo_chart.grouping = "clustered"
        emo_chart.title = "Yearly Emotion Breakdown"
        emo_chart.y_axis.title = "Average Intensity"
        emo_chart.style = 10
        emo_chart.width = 30
        emo_chart.height = 15

        for j, emo in enumerate(emotions):
            col = 7 + j
            data = Reference(ws, min_col=col, min_row=1, max_row=len(sorted_years) + 1)
            emo_chart.add_data(data, titles_from_data=True)
            emo_chart.series[j].graphicalProperties.solidFill = EMOTION_COLORS[emo]

        cats = Reference(ws, min_col=1, min_row=2, max_row=len(sorted_years) + 1)
        emo_chart.set_categories(cats)

        ws.add_chart(emo_chart, f"A{len(sorted_years) + 22}")

    _auto_width(ws)
    ws.freeze_panes = "A2"


def _create_emotions_sheet(wb: Workbook, analyzed: list[AnalyzedEpisode]):
    """Detailed emotion analysis sheet with top episodes per emotion."""
    ws = wb.create_sheet("Emotion Deep Dive")
    ws.sheet_properties.tabColor = "9370DB"

    emotions = list(EMOTION_COLORS.keys())
    row = 1

    for emo in emotions:
        # Section header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        cell = ws.cell(row=row, column=1, value=f"Top 10 Episodes — {emo.upper()}")
        cell.font = Font(bold=True, size=13, color=EMOTION_COLORS[emo])
        row += 1

        headers = ["Date", "Title", f"{emo.capitalize()} Score", "Key Words"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        _style_header(ws, row, len(headers))
        row += 1

        # Sort by this emotion score descending, take top 10
        sorted_eps = sorted(analyzed,
                            key=lambda a: a.sentiment.emotions.get(emo, 0),
                            reverse=True)[:10]

        for a in sorted_eps:
            ws.cell(row=row, column=1, value=a.episode.pub_date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=2, value=a.episode.title[:80])
            ws.cell(row=row, column=3,
                    value=round(a.sentiment.emotions.get(emo, 0), 4))
            words = a.sentiment.emotion_words.get(emo, [])
            ws.cell(row=row, column=4, value=", ".join(words[:8]))
            for c in range(1, 5):
                ws.cell(row=row, column=c).border = THIN_BORDER
            row += 1

        row += 2  # Gap between emotions

    _auto_width(ws)
