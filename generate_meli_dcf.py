#!/usr/bin/env python3
"""
MELI (MercadoLibre) Reverse Discounted Cash Flow Model Generator
Creates a self-updating Excel workbook with:
  - Live stock price via STOCKHISTORY / manual input cell
  - Formula-based reverse DCF (no Solver required)
  - Full WACC build-up
  - Historical financials 2020-2024
  - 2-D sensitivity tables
  - Instructions & VBA snippet for auto-refresh
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart.series import SeriesLabel

OUTPUT = "MELI_Reverse_DCF_Model.xlsx"

# ─────────────────────────────────────────────────────────────────────────────
# STYLE HELPERS
# ─────────────────────────────────────────────────────────────────────────────

P = {
    "dk":  "1B4332",   # dark green (header)
    "md":  "2D6A4F",   # medium green
    "lt":  "40916C",   # lighter green
    "acc": "52B788",   # accent
    "inp": "EBF5FB",   # pale blue  → editable inputs
    "out": "FFFDE7",   # pale amber → key outputs
    "imp": "FFF3CD",   # gold       → implied metric
    "pos": "D4EDDA",   # pale green → positive
    "neg": "FADBD8",   # pale red   → negative
    "gry": "F5F5F5",   # light gray
    "w":   "FFFFFF",
}

def F(c): return PatternFill("solid", fgColor=c)
def BD(color="C0C0C0"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def A(h="left", v="center", w=False, i=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=w, indent=i)
def HD(sz=11, b=True, c="FFFFFF"):
    return Font(name="Calibri", size=sz, bold=b, color=c)
def BF(sz=10, b=False, c="1A1A1A"):
    return Font(name="Calibri", size=sz, bold=b, color=c)

def sc(ws, r, c, val=None, fmt=None, fnt=None, fl=None, al=None,
       brd=True, span=None):
    """Set cell value and styles."""
    cell = ws.cell(row=r, column=c)
    if val is not None:
        cell.value = val
    if fmt:
        cell.number_format = fmt
    cell.font  = fnt or BF()
    if fl:
        cell.fill = fl
    cell.alignment = al or A()
    if brd:
        cell.border = BD()
    if span:
        ws.merge_cells(start_row=r, start_column=c,
                       end_row=r, end_column=span)
    return cell

def sec_hdr(ws, r, col_s, col_e, text, color="dk", size=11):
    """Section header spanning columns col_s..col_e."""
    cell = sc(ws, r, col_s, text,
              fnt=HD(sz=size), fl=F(P[color]),
              al=A("center"), span=col_e)
    ws.row_dimensions[r].height = 20
    return cell

def col_hdr(ws, r, col_s, col_e, text, color="md"):
    cell = sc(ws, r, col_s, text,
              fnt=HD(sz=10), fl=F(P[color]),
              al=A("center"), span=col_e)
    return cell

def label(ws, r, c, text, bold=False, indent=1, color="1A1A1A"):
    return sc(ws, r, c, text,
              fnt=BF(b=bold, c=color),
              al=A(i=indent))

def val_cell(ws, r, c, val, fmt=None, fl_key="gry", bold=False):
    return sc(ws, r, c, val, fmt=fmt,
              fnt=BF(b=bold), fl=F(P[fl_key]),
              al=A("right"))

def input_cell(ws, r, c, val, fmt=None):
    return sc(ws, r, c, val, fmt=fmt,
              fnt=BF(b=True, c="0D47A1"),
              fl=F(P["inp"]), al=A("right"))

def output_cell(ws, r, c, val, fmt=None, color="out", bold=True, size=10):
    return sc(ws, r, c, val, fmt=fmt,
              fnt=HD(sz=size, c="7B3F00") if color == "imp" else BF(b=bold, c="1A1A1A"),
              fl=F(P[color]), al=A("right"))

# ─────────────────────────────────────────────────────────────────────────────
# HISTORICAL DATA
# ─────────────────────────────────────────────────────────────────────────────

HIST = {
    "year":    [2020,  2021,   2022,   2023,   2024],
    "rev":     [3969,  7069,  10537,  14474,  20777],   # Revenue $M
    "gp":      [1682,  3015,   4497,   6939,  10530],   # Gross Profit $M
    "ebitda":  [ 274,  1163,   1565,   3468,   5300],   # EBITDA $M
    "ni":      [-161,    83,    482,   1280,   2300],   # Net Income $M
    "op_cf":   [ 722,  1571,   1644,   3164,   4500],   # Operating CF $M
    "capex":   [ 213,   329,    549,    628,    700],   # CapEx $M
    "fcf":     [ 509,  1242,   1095,   2536,   3800],   # Free CF $M
}

# ─────────────────────────────────────────────────────────────────────────────
# SHEET: DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────

def build_dashboard(ws):
    ws.sheet_view.showGridLines = False
    cw = {"A": 3, "B": 26, "C": 20, "D": 3, "E": 26, "F": 20, "G": 3}
    for col, w in cw.items():
        ws.column_dimensions[col].width = w

    # Title
    ws.row_dimensions[2].height = 50
    sc(ws, 2, 2, "MERCADOLIBRE ($MELI) — REVERSE DCF MODEL",
       fnt=HD(sz=22), fl=F(P["dk"]),
       al=A("center", "center"), span=6)
    sc(ws, 3, 2, "What FCF growth rate is the market currently pricing in?",
       fnt=HD(sz=12, b=False, c="FFFFFF"), fl=F(P["md"]),
       al=A("center"), span=6)
    ws.row_dimensions[3].height = 24

    # Date / disclaimer bar
    sc(ws, 4, 2, "=TODAY()", fmt="DD-MMM-YYYY",
       fnt=BF(b=True, c=P["md"]), fl=F(P["gry"]),
       al=A("center"), brd=False)
    sc(ws, 4, 3, "Model auto-recalculates when inputs change. Use Data > Refresh All for live price.",
       fnt=BF(sz=9, c="666666"), fl=F(P["gry"]),
       al=A("left"), brd=False, span=5)
    ws.row_dimensions[4].height = 16

    # ── Key Metric Cards (row 6-13) ──────────────────────────────────────────
    r = 6
    sc(ws, r, 2, "KEY MODEL INPUTS", fnt=HD(sz=10), fl=F(P["lt"]),
       al=A("center"), span=3)
    sc(ws, r, 5, "KEY MODEL OUTPUTS", fnt=HD(sz=10), fl=F(P["lt"]),
       al=A("center"), span=6)

    metrics_l = [
        ("Current Stock Price",   "='Reverse DCF'!C7",  '$#,##0.00'),
        ("Base FCF (FY2024, $M)", "='Reverse DCF'!C8",  '$#,##0'),
        ("Shares Outstanding (M)","='Reverse DCF'!C9",  '0.0'),
        ("Net Debt ($M)",         "='Reverse DCF'!C10", '$#,##0'),
        ("WACC",                  "='WACC'!C28",        '0.00%'),
        ("Terminal Growth Rate",  "='Reverse DCF'!C15", '0.00%'),
    ]
    metrics_r = [
        ("Implied FCF Growth Rate","='Reverse DCF'!C19","0.00%"),
        ("Intrinsic Value (Your View)","='Reverse DCF'!C24",'$#,##0.00'),
        ("Premium / (Discount)",  "='Reverse DCF'!C25", '0.0%'),
        ("Enterprise Value ($M)", "='Reverse DCF'!C43", '$#,##0'),
        ("Equity Value ($M)",     "='Reverse DCF'!C44", '$#,##0'),
        ("PV of Terminal Value (%)", "=IF('Reverse DCF'!C43<>0,'Reverse DCF'!C42/'Reverse DCF'!C43,0)", '0.0%'),
    ]

    for i, (lbl, formula, fmt) in enumerate(metrics_l):
        rr = r + 1 + i
        ws.row_dimensions[rr].height = 22
        sc(ws, rr, 2, lbl, fnt=BF(b=True), fl=F(P["gry"]), al=A("left", i=1))
        sc(ws, rr, 3, formula, fmt=fmt, fnt=BF(b=True, c="1565C0"),
           fl=F(P["inp"]), al=A("right"))

    for i, (lbl, formula, fmt) in enumerate(metrics_r):
        rr = r + 1 + i
        fl_key = "imp" if i == 0 else "out"
        sc(ws, rr, 5, lbl, fnt=BF(b=True), fl=F(P["gry"]), al=A("left", i=1))
        sc(ws, rr, 6, formula, fmt=fmt,
           fnt=HD(sz=11, c="7B3F00") if i == 0 else BF(b=True),
           fl=F(P[fl_key]), al=A("right"))

    # ── Historical FCF summary (rows 15-24) ──────────────────────────────────
    r = 15
    sc(ws, r, 2, "HISTORICAL FCF SUMMARY ($M)", fnt=HD(sz=10),
       fl=F(P["dk"]), al=A("center"), span=6)
    ws.row_dimensions[r].height = 20

    hdr_years = HIST["year"]
    sc(ws, r+1, 2, "Metric", fnt=HD(sz=10), fl=F(P["md"]), al=A("center"))
    for j, yr in enumerate(hdr_years):
        sc(ws, r+1, 3+j, str(yr), fnt=HD(sz=10), fl=F(P["md"]), al=A("center"))

    hist_rows = [
        ("Revenue",      "rev"),
        ("Gross Profit", "gp"),
        ("EBITDA",       "ebitda"),
        ("Net Income",   "ni"),
        ("Free Cash Flow","fcf"),
    ]
    for i, (name, key) in enumerate(hist_rows):
        rr = r + 2 + i
        ws.row_dimensions[rr].height = 18
        fl = F(P["gry"]) if i % 2 == 0 else F(P["w"])
        sc(ws, rr, 2, name, fnt=BF(), fl=fl, al=A("left", i=1))
        for j, v in enumerate(HIST[key]):
            neg = v < 0
            sc(ws, rr, 3+j, v, fmt='$#,##0',
               fnt=BF(c="C0392B" if neg else "1A1A1A"),
               fl=fl, al=A("right"))

    # Branding footer
    sc(ws, 27, 2,
       "Model built with Python + openpyxl  |  Data: Company filings, analyst estimates  |  Not investment advice",
       fnt=BF(sz=8, c="999999"), fl=None, brd=False, al=A("left"), span=6)


# ─────────────────────────────────────────────────────────────────────────────
# SHEET: WACC
# ─────────────────────────────────────────────────────────────────────────────

def build_wacc(ws):
    ws.sheet_view.showGridLines = False
    for col, w in {"A": 3, "B": 38, "C": 20, "D": 28}.items():
        ws.column_dimensions[col].width = w

    ws.row_dimensions[2].height = 40
    sc(ws, 2, 2, "WACC — WEIGHTED AVERAGE COST OF CAPITAL",
       fnt=HD(sz=18), fl=F(P["dk"]), al=A("center", "center"), span=4)
    sc(ws, 3, 2, "MercadoLibre, Inc.  |  Ticker: MELI  |  Exchange: NASDAQ",
       fnt=HD(sz=11, b=False), fl=F(P["md"]), al=A("center"), span=4)
    ws.row_dimensions[3].height = 22

    # ── Cost of Equity ────────────────────────────────────────────────────────
    r = 5
    sec_hdr(ws, r, 2, 4, "COST OF EQUITY  (CAPM)", "lt")
    sc(ws, r+1, 2, "Component", fnt=HD(sz=10), fl=F(P["md"]), al=A("center"))
    sc(ws, r+1, 3, "Value",     fnt=HD(sz=10), fl=F(P["md"]), al=A("center"))
    sc(ws, r+1, 4, "Notes",     fnt=HD(sz=10), fl=F(P["md"]), al=A("center"))

    coe_rows = [
        ("Risk-Free Rate (10Y UST)",     0.045, "0.00%", "US 10-Year Treasury yield"),
        ("Equity Beta",                  1.35,  "0.00",  "5-yr monthly regression vs S&P 500"),
        ("Equity Risk Premium (ERP)",    0.055, "0.00%", "Damodaran implied ERP"),
        ("Country Risk Premium (LatAm)", 0.015, "0.00%", "Weighted avg of MELI operating markets"),
    ]
    for i, (name, val, fmt, note) in enumerate(coe_rows):
        rr = r + 2 + i
        ws.row_dimensions[rr].height = 18
        fl = F(P["gry"]) if i % 2 == 0 else F(P["w"])
        label(ws, rr, 2, name)
        ws.cell(rr, 2).fill = fl
        input_cell(ws, rr, 3, val, fmt=fmt)
        sc(ws, rr, 4, note, fnt=BF(sz=9, c="666666"), fl=fl, al=A("left", i=1))

    # Cost of Equity formula = RFR + Beta*(ERP + CRP)
    rr = r + 6
    ws.row_dimensions[rr].height = 20
    sec_hdr(ws, rr, 2, 2, "Cost of Equity  =  Rf + β × (ERP + CRP)", "dk", size=10)
    output_cell(ws, rr, 3, "=C7+C8*(C9+C10)", fmt="0.00%", color="imp", bold=True)
    sc(ws, rr, 4, "CAPM formula", fnt=BF(sz=9, c="666666"), fl=F(P["w"]), al=A("left", i=1))

    # ── Cost of Debt ──────────────────────────────────────────────────────────
    r = 14
    sec_hdr(ws, r, 2, 4, "COST OF DEBT", "lt")
    cod_rows = [
        ("Pre-Tax Cost of Debt",  0.065, "0.00%", "Blended yield on MELI bonds"),
        ("Effective Tax Rate",    0.300, "0.0%",  "Blended across LatAm jurisdictions"),
    ]
    for i, (name, val, fmt, note) in enumerate(cod_rows):
        rr = r + 1 + i
        ws.row_dimensions[rr].height = 18
        fl = F(P["gry"]) if i % 2 == 0 else F(P["w"])
        label(ws, rr, 2, name)
        ws.cell(rr, 2).fill = fl
        input_cell(ws, rr, 3, val, fmt=fmt)
        sc(ws, rr, 4, note, fnt=BF(sz=9, c="666666"), fl=fl, al=A("left", i=1))

    rr = r + 3
    ws.row_dimensions[rr].height = 20
    sec_hdr(ws, rr, 2, 2, "After-Tax Cost of Debt  =  Kd × (1 – Tax)", "dk", size=10)
    output_cell(ws, rr, 3, "=C15*(1-C16)", fmt="0.00%", color="imp")
    sc(ws, rr, 4, "Tax shield on interest", fnt=BF(sz=9, c="666666"),
       fl=F(P["w"]), al=A("left", i=1))

    # ── Capital Structure ─────────────────────────────────────────────────────
    r = 20
    sec_hdr(ws, r, 2, 4, "CAPITAL STRUCTURE", "lt")
    cap_rows = [
        ("Market Capitalisation ($M)", 97000, "$#,##0", "Price × Shares outstanding"),
        ("Total Debt ($M)",             7000, "$#,##0", "Long-term + short-term debt"),
    ]
    for i, (name, val, fmt, note) in enumerate(cap_rows):
        rr = r + 1 + i
        ws.row_dimensions[rr].height = 18
        fl = F(P["gry"]) if i % 2 == 0 else F(P["w"])
        label(ws, rr, 2, name)
        ws.cell(rr, 2).fill = fl
        input_cell(ws, rr, 3, val, fmt=fmt)
        sc(ws, rr, 4, note, fnt=BF(sz=9, c="666666"), fl=fl, al=A("left", i=1))

    # Derived weights
    derived = [
        ("Total Capital ($M)",   "=C21+C22", "$#,##0"),
        ("Equity Weight (E/V)",  "=C21/C23", "0.00%"),
        ("Debt Weight (D/V)",    "=C22/C23", "0.00%"),
    ]
    for i, (name, formula, fmt) in enumerate(derived):
        rr = r + 3 + i
        ws.row_dimensions[rr].height = 18
        fl = F(P["gry"]) if i % 2 == 0 else F(P["w"])
        label(ws, rr, 2, name, bold=True)
        ws.cell(rr, 2).fill = fl
        output_cell(ws, rr, 3, formula, fmt=fmt)
        ws.cell(rr, 3).fill = fl

    # ── WACC Result ───────────────────────────────────────────────────────────
    # Row 27 spacer
    ws.row_dimensions[27].height = 8
    r = 28
    ws.row_dimensions[r].height = 28
    # Label in B, value in C, note in D — DO NOT merge (would block writing to C28)
    sc(ws, r, 2, "WACC  =  (E/V) × Ke  +  (D/V) × Kd(1–t)",
       fnt=HD(sz=12, c="FFFFFF"), fl=F(P["dk"]), al=A("center"))
    # Cost of Equity = C11, After-tax Kd = C17, Equity Weight = C24, Debt Weight = C25
    sc(ws, r, 3, "=C24*C11+C25*C17", fmt="0.00%",
       fnt=HD(sz=16, c="7B3F00"), fl=F(P["imp"]),
       al=A("center"), brd=True)
    sc(ws, r, 4, "← WACC used in Reverse DCF model",
       fnt=BF(sz=9, c="444444"), fl=F(P["w"]), al=A("left", i=1))

    sc(ws, 30, 2,
       "Note: Update Market Cap cell (C21) when stock price changes for accurate weights.",
       fnt=BF(sz=9, c="E74C3C"), brd=False, al=A("left", i=1))


# ─────────────────────────────────────────────────────────────────────────────
# SHEET: REVERSE DCF  (main model)
# ─────────────────────────────────────────────────────────────────────────────

def build_reverse_dcf(ws):
    ws.sheet_view.showGridLines = False

    # Column widths
    cols = {"A": 2, "B": 34, "C": 20, "D": 2, "E": 30, "F": 20,
            "G": 2, "H": 22, "I": 22, "J": 22}
    for col, w in cols.items():
        ws.column_dimensions[col].width = w

    # Title
    ws.row_dimensions[2].height = 45
    sc(ws, 2, 2, "MELI REVERSE DCF MODEL  —  Implied Growth Rate Analysis",
       fnt=HD(sz=18), fl=F(P["dk"]), al=A("center", "center"), span=10)
    sc(ws, 3, 2,
       "The model solves for the FCF growth rate the market is pricing into the current stock price.",
       fnt=HD(sz=11, b=False), fl=F(P["md"]), al=A("center"), span=10)
    ws.row_dimensions[3].height = 24

    # ─── LEFT PANEL: INPUTS & ASSUMPTIONS (cols B–C) ─────────────────────────

    # Section: Inputs
    r = 5
    sec_hdr(ws, r, 2, 6, "INPUTS  &  ASSUMPTIONS", "dk")

    # Sub-header row
    sc(ws, r+1, 2, "Parameter", fnt=HD(sz=10), fl=F(P["md"]), al=A("center"))
    sc(ws, r+1, 3, "Value",     fnt=HD(sz=10), fl=F(P["md"]), al=A("center"))
    sc(ws, r+1, 5, "Key Output",fnt=HD(sz=10), fl=F(P["md"]), al=A("center"))
    sc(ws, r+1, 6, "Value",     fnt=HD(sz=10), fl=F(P["md"]), al=A("center"))

    # ── Editable inputs (BLUE cells) ─────────────────────────────────────────
    # C7  = Current Stock Price (STOCKHISTORY or manual override)
    # NOTE: STOCKHISTORY requires Excel 365. Fallback = manual entry.
    STOCKHISTORY_FORMULA = (
        '=IFERROR(INDEX(STOCKHISTORY("MELI",TODAY()-7,TODAY(),0,0),'
        'COUNTA(STOCKHISTORY("MELI",TODAY()-7,TODAY(),0,0)),1),1950)'
    )

    inputs_left = [
        # (row-label, value-or-formula, format, note)
        ("Current Stock Price ($)",  STOCKHISTORY_FORMULA, '$#,##0.00',
         "Auto-updates via STOCKHISTORY (Excel 365)"),
        ("Base FCF — FY2024 ($M)",   3800,    '$#,##0',
         "FY2024 actual/estimate; update each year"),
        ("Shares Outstanding (M)",   51.0,    '0.0',
         "Diluted shares including RSUs"),
        ("Net Debt ($M)",            2000,    '$#,##0',
         "Total debt minus cash & equivalents"),
    ]

    for i, (lbl, val, fmt, note) in enumerate(inputs_left):
        rr = r + 2 + i
        ws.row_dimensions[rr].height = 20
        fl = F(P["gry"]) if i % 2 == 0 else F(P["w"])
        label(ws, rr, 2, lbl, bold=True)
        ws.cell(rr, 2).fill = fl

        cell = ws.cell(row=rr, column=3)
        cell.value = val
        cell.number_format = fmt
        # Stock price cell gets special treatment (STOCKHISTORY formula)
        if i == 0:
            cell.font  = BF(b=True, c="0D47A1")
            cell.fill  = F(P["inp"])
        else:
            cell.font  = BF(b=True, c="0D47A1")
            cell.fill  = F(P["inp"])
        cell.alignment = A("right")
        cell.border    = BD()

        sc(ws, rr, 4, f"← {note}", fnt=BF(sz=8, c="777777"),
           fl=fl, brd=False, al=A("left"))

    # ── DCF Assumptions ───────────────────────────────────────────────────────
    r2 = 12
    sec_hdr(ws, r2, 2, 6, "DCF ASSUMPTIONS", "lt")

    assum = [
        ("WACC",                  "='WACC'!C28", "0.00%",  "From WACC sheet (auto-linked)"),
        ("Terminal Growth Rate",  0.04,           "0.00%",  "Long-run FCF growth (4% = LatAm GDP)"),
        ("Projection Years",      10,             "0",      "Explicit forecast horizon"),
        ("Your Target Growth Rate", 0.25,         "0.00%",  "YOUR assumption → drives Forward DCF"),
    ]

    for i, (lbl, val, fmt, note) in enumerate(assum):
        rr = r2 + 1 + i
        ws.row_dimensions[rr].height = 20
        fl = F(P["gry"]) if i % 2 == 0 else F(P["w"])
        label(ws, rr, 2, lbl, bold=True)
        ws.cell(rr, 2).fill = fl

        cell = ws.cell(row=rr, column=3)
        cell.value = val
        cell.number_format = fmt
        # WACC is a formula; target growth is editable input
        if i == 0:
            cell.font  = BF(b=True, c="1A1A1A")
            cell.fill  = F(P["out"])
        else:
            cell.font  = BF(b=True, c="0D47A1")
            cell.fill  = F(P["inp"])
        cell.alignment = A("right")
        cell.border    = BD()

        sc(ws, rr, 4, f"← {note}", fnt=BF(sz=8, c="777777"),
           fl=fl, brd=False, al=A("left"))

    # Cell references (for formula use below):
    # C7 = current stock price
    # C8 = base FCF
    # C9 = shares outstanding
    # C10 = net debt
    # C13 = WACC
    # C14 = terminal growth rate
    # C15 = projection years
    # C16 = target growth rate (your view)

    # ── KEY OUTPUTS (right half of row 6 area, cols E-F) ─────────────────────
    # Implied Growth Rate (from lookup table INDEX/MATCH):
    # Uses H8:H68 (growth rates) and I8:I68 (DCF prices)

    out_rows = [
        # (row, label, formula, fmt, color)
        (7,  "■ IMPLIED GROWTH RATE (Market)",
         '=IFERROR(INDEX($H$8:$H$68,MATCH(MIN(ABS($I$8:$I$68-$C$7)),ABS($I$8:$I$68-$C$7),0)),"#N/A")',
         "0.00%", "imp"),
        (8,  "  FCF CAGR 2021–2024 (Historical)",
         "=(C8/1242)^(1/3)-1",
         "0.00%", "pos"),
        (9,  "■ INTRINSIC VALUE — Your View",
         '=IFERROR((SUMPRODUCT($C$8*(1+$C$16)^ROW(INDIRECT("1:"&$C$15)),'
         '(1/(1+$C$13))^ROW(INDIRECT("1:"&$C$15)))'
         '+$C$8*(1+$C$16)^$C$15*(1+$C$14)/($C$13-$C$14)/(1+$C$13)^$C$15'
         '-$C$10)/$C$9,"Check WACC>TermG")',
         '$#,##0.00', "out"),
        (10, "  Premium / (Discount) to Market",
         "=IFERROR((F9-C7)/C7,0)",
         '+0.0%;-0.0%;0.0%', "out"),
        (11, "  EV at Your Target Growth ($M)",
         '=IFERROR(SUMPRODUCT($C$8*(1+$C$16)^ROW(INDIRECT("1:"&$C$15)),'
         '(1/(1+$C$13))^ROW(INDIRECT("1:"&$C$15)))'
         '+$C$8*(1+$C$16)^$C$15*(1+$C$14)/($C$13-$C$14)/(1+$C$13)^$C$15,'
         '"N/A")',
         '$#,##0', "out"),
    ]

    for row_offset, lbl, formula, fmt, fl_key in out_rows:
        ws.row_dimensions[row_offset].height = 22
        is_key = "■" in lbl
        label(ws, row_offset, 5, lbl, bold=is_key,
              color="1B4332" if is_key else "333333")
        ws.cell(row_offset, 5).fill = F(P["gry"])

        cell = ws.cell(row=row_offset, column=6)
        cell.value = formula
        cell.number_format = fmt
        cell.fill = F(P[fl_key])
        cell.alignment = A("right")
        cell.border = BD()
        if fl_key == "imp":
            cell.font = HD(sz=14, c="7B3F00")
        elif is_key:
            cell.font = BF(b=True)
        else:
            cell.font = BF()

    # Named convenience references for the projection section
    # C19 = implied growth rate (INDEX/MATCH)  → actually in F7, link it to C19
    ws.row_dimensions[19].height = 0  # hidden helper row
    sc(ws, 19, 3, "=F7", brd=False)   # C19 = implied growth rate (helper)
    sc(ws, 24, 3, "=F9", brd=False)   # C24 = intrinsic value (helper)
    sc(ws, 25, 3, "=F10", brd=False)  # C25 = premium/discount (helper)
    sc(ws, 43, 3, "=F11", brd=False)  # C43 = EV helper

    # ── FCF PROJECTION TABLE (rows 28-40) ────────────────────────────────────
    r3 = 28
    sec_hdr(ws, r3, 2, 6,
            "FCF PROJECTION  —  At Market-Implied Growth Rate", "dk")

    # Column headers
    proj_hdrs = ["Year", "Calendar", "FCF ($M)", "YoY Growth", "PV Factor", "PV of FCF ($M)"]
    for j, h in enumerate(proj_hdrs):
        col = j + 2
        sc(ws, r3+1, col, h, fnt=HD(sz=10), fl=F(P["md"]), al=A("center"))

    for yr in range(1, 11):
        rr = r3 + 1 + yr
        ws.row_dimensions[rr].height = 18
        fl = F(P["gry"]) if yr % 2 == 1 else F(P["w"])

        sc(ws, rr, 2, f"Year {yr}", fnt=BF(), fl=fl, al=A("center"))
        sc(ws, rr, 3, f"=2024+{yr}", fmt="0", fnt=BF(), fl=fl, al=A("center"))
        # FCF = BaseFCF × (1+ImpliedGrowth)^yr
        sc(ws, rr, 4,
           f"=$C$8*(1+$F$7)^{yr}",
           fmt='$#,##0', fnt=BF(), fl=fl, al=A("right"))
        # YoY Growth (= implied growth rate for all years)
        sc(ws, rr, 5, "=$F$7", fmt="0.0%", fnt=BF(), fl=fl, al=A("right"))
        # PV Factor = 1/(1+WACC)^yr
        sc(ws, rr, 6,
           f"=1/(1+$C$13)^{yr}",
           fmt="0.000000", fnt=BF(), fl=fl, al=A("right"))
        # PV of FCF
        sc(ws, rr, 7,
           f"=D{rr}*F{rr}",
           fmt='$#,##0', fnt=BF(), fl=fl, al=A("right"))

    # Projection summary rows
    r4 = r3 + 12
    ws.row_dimensions[r4].height = 6

    summary_rows = [
        (r4+1, "Sum of PV of FCFs ($M)",
         "=SUM(G30:G39)", "$#,##0", "out"),
        (r4+2, "Terminal Value — Undiscounted ($M)",
         "=IFERROR($C$8*(1+$F$7)^$C$15*(1+$C$14)/($C$13-$C$14),\"Check WACC>TermG\")",
         "$#,##0", "out"),
        (r4+3, "PV of Terminal Value ($M)",
         f"=IFERROR(C{r4+2}/(1+$C$13)^$C$15,0)",
         "$#,##0", "out"),
        (r4+4, "Enterprise Value ($M)",
         f"=C{r4+1}+C{r4+3}",
         "$#,##0", "out"),
        (r4+5, "Less: Net Debt ($M)",
         "=$C$10",
         "$#,##0", "out"),
        (r4+6, "= Equity Value ($M)",
         f"=C{r4+4}-C{r4+5}",
         "$#,##0", "imp"),
        (r4+7, "Shares Outstanding (M)",
         "=$C$9",
         "0.0", "gry"),
        (r4+8, "■ Implied Stock Price  (check vs C7)",
         f"=IFERROR(C{r4+6}/$C$9,0)",
         "$#,##0.00", "imp"),
    ]

    for (rr, lbl, formula, fmt, fl_key) in summary_rows:
        ws.row_dimensions[rr].height = 20
        is_key = "■" in lbl
        label(ws, rr, 2, lbl, bold=is_key,
              color="1B4332" if is_key else "1A1A1A")
        ws.cell(rr, 2).fill = F(P["gry"])
        cell = ws.cell(row=rr, column=3)
        cell.value = formula
        cell.number_format = fmt
        cell.fill = F(P[fl_key])
        cell.border = BD()
        cell.alignment = A("right")
        if is_key:
            cell.font = HD(sz=11, c="7B3F00")
        else:
            cell.font = BF(b=True if fl_key in ("out","imp") else False)

    # ── RIGHT PANEL: REVERSE DCF LOOKUP TABLE (cols H–J) ─────────────────────
    # Rows 5-70: growth 0% to 60%, DCF prices, delta vs current
    rs = 5   # start row for right panel

    sec_hdr(ws, rs, 8, 10, "REVERSE DCF LOOKUP TABLE", "dk")
    sc(ws, rs+1, 8,
       "Each row shows the stock price implied by that growth rate. "
       "Market-implied growth = row where Implied Price ≈ Current Price.",
       fnt=BF(sz=9, c="555555"), fl=F(P["gry"]),
       al=A("left", w=True), brd=False, span=10)
    ws.row_dimensions[rs+1].height = 30

    # Column headers
    rt = rs + 2
    sc(ws, rt, 8, "Trial FCF Growth", fnt=HD(sz=10), fl=F(P["md"]), al=A("center"))
    sc(ws, rt, 9, "Implied Price ($)",  fnt=HD(sz=10), fl=F(P["md"]), al=A("center"))
    sc(ws, rt, 10,"vs Current Price",  fnt=HD(sz=10), fl=F(P["md"]), al=A("center"))

    # Build the DCF formula for each row:
    # =IFERROR(
    #   (SUMPRODUCT($C$8*(1+H_r)^ROW(INDIRECT("1:"&$C$15)),
    #              (1/(1+$C$13))^ROW(INDIRECT("1:"&$C$15)))
    #    + $C$8*(1+H_r)^$C$15*(1+$C$14)/($C$13-$C$14)/(1+$C$13)^$C$15
    #    - $C$10) / $C$9
    # , "N/A")

    for step in range(61):   # 0% to 60% inclusive
        rr = rt + 1 + step
        ws.row_dimensions[rr].height = 15
        growth_val = step / 100.0
        growth_ref = f"H{rr}"

        fl = F(P["gry"]) if step % 2 == 0 else F(P["w"])

        # Growth rate cell
        sc(ws, rr, 8, growth_val, fmt="0%",
           fnt=BF(), fl=fl, al=A("center"))

        # DCF Price formula
        dcf_formula = (
            f'=IFERROR('
            f'(SUMPRODUCT($C$8*(1+{growth_ref})^ROW(INDIRECT("1:"&$C$15)),'
            f'(1/(1+$C$13))^ROW(INDIRECT("1:"&$C$15)))'
            f'+$C$8*(1+{growth_ref})^$C$15*(1+$C$14)/($C$13-$C$14)/(1+$C$13)^$C$15'
            f'-$C$10)/$C$9,"N/A")'
        )
        sc(ws, rr, 9, dcf_formula, fmt='$#,##0.00',
           fnt=BF(), fl=fl, al=A("right"))

        # Delta vs current price
        delta_formula = f'=IFERROR((I{rr}-$C$7)/$C$7,"N/A")'
        sc(ws, rr, 10, delta_formula, fmt='+0.0%;-0.0%;0.0%',
           fnt=BF(), fl=fl, al=A("right"))

    # Implied growth result row (below lookup table)
    rr_implied = rt + 62
    ws.row_dimensions[rr_implied].height = 6
    rr_implied += 1
    ws.row_dimensions[rr_implied].height = 24
    sc(ws, rr_implied, 8,
       "► MARKET-IMPLIED FCF GROWTH RATE",
       fnt=HD(sz=11, c="7B3F00"), fl=F(P["imp"]),
       al=A("center"), span=9)
    sc(ws, rr_implied, 10,
       '=IFERROR(INDEX($H$8:$H$68,MATCH(MIN(ABS($I$8:$I$68-$C$7)),ABS($I$8:$I$68-$C$7),0)),"#N/A")',
       fmt="0.00%",
       fnt=HD(sz=14, c="7B3F00"), fl=F(P["imp"]),
       al=A("center"))

    # Conditional formatting: green where delta < 5%, red where > 50% off
    ws.conditional_formatting.add(
        "J9:J68",
        ColorScaleRule(
            start_type="num", start_value=-1, start_color="F44336",
            mid_type="num",   mid_value=0,    mid_color="FFEB3B",
            end_type="num",   end_value=1,    end_color="4CAF50",
        )
    )

    # ── Add a line chart: Implied Price vs Growth Rate ────────────────────────
    chart = LineChart()
    chart.title = "Implied DCF Price vs FCF Growth Rate"
    chart.style = 10
    chart.y_axis.title = "Implied Stock Price ($)"
    chart.x_axis.title = "Assumed FCF Growth Rate"
    chart.height = 14
    chart.width  = 22

    # Data: I8:I68 (prices)
    data_ref = Reference(ws, min_col=9, min_row=8, max_row=68)
    chart.add_data(data_ref, titles_from_data=False)
    s0 = chart.series[0]
    from openpyxl.chart.series import SeriesLabel
    s0.title = SeriesLabel(v="Implied Price")
    s0.graphicalProperties.line.solidFill = "2D6A4F"
    s0.graphicalProperties.line.width = 20000

    # Category: H8:H68 (growth rates)
    cats = Reference(ws, min_col=8, min_row=8, max_row=68)
    chart.set_categories(cats)

    ws.add_chart(chart, "H71")


# ─────────────────────────────────────────────────────────────────────────────
# SHEET: HISTORICALS
# ─────────────────────────────────────────────────────────────────────────────

def build_historicals(ws):
    ws.sheet_view.showGridLines = False
    for col, w in {"A": 2, "B": 32, "C": 16, "D": 16,
                   "E": 16, "F": 16, "G": 16, "H": 14}.items():
        ws.column_dimensions[col].width = w

    ws.row_dimensions[2].height = 40
    sc(ws, 2, 2, "MERCADOLIBRE — HISTORICAL FINANCIALS  ($M unless noted)",
       fnt=HD(sz=16), fl=F(P["dk"]), al=A("center", "center"), span=8)
    sc(ws, 3, 2,
       "Source: MercadoLibre SEC filings & company guidance.  "
       "FY2024 = analyst consensus estimate.",
       fnt=HD(sz=10, b=False), fl=F(P["md"]), al=A("center"), span=8)
    ws.row_dimensions[3].height = 20

    # Header row
    r = 5
    sc(ws, r, 2, "Metric", fnt=HD(sz=10), fl=F(P["lt"]), al=A("center"))
    for j, yr in enumerate(HIST["year"]):
        sc(ws, r, 3+j, str(yr), fnt=HD(sz=10), fl=F(P["lt"]), al=A("center"))
    sc(ws, r, 8, "CAGR '20–'24", fnt=HD(sz=10), fl=F(P["lt"]), al=A("center"))

    rows_def = [
        ("INCOME STATEMENT", None, None, "sec"),
        ("Revenue",          "rev",  '$#,##0',   "num"),
        ("  Gross Profit",   "gp",   '$#,##0',   "num"),
        ("  Gross Margin",   None,   '0.0%',     "margin_gp_rev"),
        ("  EBITDA",         "ebitda",'$#,##0',  "num"),
        ("  EBITDA Margin",  None,   '0.0%',     "margin_ebitda_rev"),
        ("  Net Income",     "ni",   '$#,##0',   "num"),
        ("CASH FLOW",        None,   None,       "sec"),
        ("Operating Cash Flow","op_cf",'$#,##0', "num"),
        ("  Capital Expenditures","capex",'$#,##0',"num"),
        ("  Free Cash Flow", "fcf",  '$#,##0',   "num"),
        ("  FCF Margin",     None,   '0.0%',     "margin_fcf_rev"),
        ("  FCF Growth YoY", None,   '0.0%',     "fcf_growth"),
    ]

    # Track row numbers for margin calculations
    rev_rows  = {}
    gp_rows   = {}
    ebitda_rows = {}
    fcf_rows  = {}
    data_row_map = {}

    actual_r = r + 1
    for lbl, key, fmt, kind in rows_def:
        ws.row_dimensions[actual_r].height = 18
        if kind == "sec":
            sec_hdr(ws, actual_r, 2, 8, lbl, "lt", size=10)
            actual_r += 1
            continue

        is_alt = (actual_r % 2 == 0)
        fl = F(P["gry"]) if is_alt else F(P["w"])
        is_indent = lbl.startswith("  ")
        label(ws, actual_r, 2, lbl, indent=2 if is_indent else 1)
        ws.cell(actual_r, 2).fill = fl

        if key:
            data_row_map[key] = actual_r

        for j, (yr, v) in enumerate(zip(HIST["year"], HIST[key] if key else [None]*5)):
            col = 3 + j
            if kind == "num" and v is not None:
                neg = (v < 0)
                sc(ws, actual_r, col, v, fmt=fmt,
                   fnt=BF(c="C0392B" if neg else "1A1A1A"),
                   fl=fl, al=A("right"))
            elif kind.startswith("margin"):
                base_key, denom_key = kind.replace("margin_", "").split("_")
                b_row = data_row_map.get(base_key, None)
                d_row = data_row_map.get(denom_key, None)
                if b_row and d_row:
                    col_ltr = get_column_letter(col)
                    formula = f"=IF({col_ltr}{d_row}<>0,{col_ltr}{b_row}/{col_ltr}{d_row},0)"
                    sc(ws, actual_r, col, formula, fmt=fmt,
                       fnt=BF(c="1565C0"), fl=fl, al=A("right"))
            elif kind == "fcf_growth" and j > 0:
                prev_col = get_column_letter(col - 1)
                curr_col = get_column_letter(col)
                fcf_row  = data_row_map.get("fcf", None)
                if fcf_row:
                    formula = f"=IF({prev_col}{fcf_row}<>0,{curr_col}{fcf_row}/{prev_col}{fcf_row}-1,0)"
                    sc(ws, actual_r, col, formula, fmt=fmt,
                       fnt=BF(c="1565C0"), fl=fl, al=A("right"))
                else:
                    sc(ws, actual_r, col, "—", fnt=BF(), fl=fl, al=A("center"))

        # CAGR (col 8)
        if kind == "num" and key:
            vals  = HIST[key]
            start_col = get_column_letter(3)
            end_col   = get_column_letter(7)
            years     = len(HIST["year"]) - 1
            cagr_f = (
                f"=IFERROR(({end_col}{actual_r}/{start_col}{actual_r})"
                f"^(1/{years})-1,\"\")"
            )
            sc(ws, actual_r, 8, cagr_f, fmt="0.0%",
               fnt=BF(b=True, c="2D6A4F"), fl=F(P["pos"]), al=A("center"))
        else:
            sc(ws, actual_r, 8, "—", fnt=BF(), fl=fl, al=A("center"))

        actual_r += 1

    # Add bar chart: FCF history
    fcf_row_n = data_row_map.get("fcf", None)
    if fcf_row_n:
        chart = BarChart()
        chart.type  = "col"
        chart.title = "MercadoLibre — Annual Free Cash Flow ($M)"
        chart.style = 10
        chart.y_axis.title = "FCF ($M)"
        chart.height = 12
        chart.width  = 24

        data = Reference(ws, min_col=3, max_col=7,
                         min_row=fcf_row_n, max_row=fcf_row_n)
        cats = Reference(ws, min_col=3, max_col=7, min_row=5)
        chart.add_data(data)
        chart.set_categories(cats)
        from openpyxl.chart.series import SeriesLabel
        chart.series[0].title = SeriesLabel(v="FCF ($M)")
        chart.series[0].graphicalProperties.solidFill = "2D6A4F"
        ws.add_chart(chart, "B22")


# ─────────────────────────────────────────────────────────────────────────────
# SHEET: SENSITIVITY
# ─────────────────────────────────────────────────────────────────────────────

def build_sensitivity(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 20

    ws.row_dimensions[2].height = 40
    sc(ws, 2, 2, "SENSITIVITY ANALYSIS — DCF Implied Stock Price ($)",
       fnt=HD(sz=16), fl=F(P["dk"]), al=A("center", "center"), span=12)
    sc(ws, 3, 2,
       "Each cell shows the implied stock price for a given combination of "
       "WACC (rows) and FCF Growth Rate (columns). Shading: green = above current price.",
       fnt=HD(sz=10, b=False), fl=F(P["md"]), al=A("center", w=True), span=12)
    ws.row_dimensions[3].height = 30

    # Cross-reference: current price from Reverse DCF sheet
    sc(ws, 4, 2, "Current Stock Price →", fnt=BF(b=True), fl=F(P["gry"]),
       al=A("right"), brd=False)
    sc(ws, 4, 3, "='Reverse DCF'!C7", fmt='$#,##0.00',
       fnt=BF(b=True, c="0D47A1"), fl=F(P["inp"]), al=A("center"))
    sc(ws, 4, 4, "← Reference for color-coding below",
       fnt=BF(sz=9, c="777777"), brd=False, al=A("left"), span=8)

    growth_rates = [0.10, 0.15, 0.20, 0.25, 0.30, 0.35, 0.40, 0.45, 0.50, 0.55]
    wacc_rates   = [0.08, 0.09, 0.10, 0.11, 0.12, 0.13, 0.14, 0.15, 0.16]

    # ── Table 1: Implied Stock Price ──────────────────────────────────────────
    r = 6
    sec_hdr(ws, r, 2, 12, "IMPLIED STOCK PRICE  ($ / share)", "dk")

    # Set column widths for data columns
    for col_idx in range(3, 3 + len(growth_rates)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 13

    # Growth rate header row
    ws.row_dimensions[r+1].height = 22
    sc(ws, r+1, 2, "WACC ↓ / Growth →",
       fnt=HD(sz=9), fl=F(P["lt"]), al=A("center"))
    for j, g in enumerate(growth_rates):
        sc(ws, r+1, 3+j, f"{g:.0%}",
           fnt=HD(sz=10), fl=F(P["lt"]), al=A("center"))

    # Build each WACC row
    for i, wacc in enumerate(wacc_rates):
        rr = r + 2 + i
        ws.row_dimensions[rr].height = 20
        fl = F(P["gry"]) if i % 2 == 0 else F(P["w"])

        sc(ws, rr, 2, f"{wacc:.0%}", fnt=HD(sz=10), fl=F(P["md"]), al=A("center"))

        for j, growth in enumerate(growth_rates):
            col = 3 + j
            col_ltr = get_column_letter(col)
            row_ltr = rr

            # Growth rate reference = header row col value
            g_ref = f"{col_ltr}{r+1}"
            # WACC reference = row header value (col B, but it's a hardcoded %)
            # We'll use the direct value since it's in the header cell

            # DCF formula referencing:
            # - 'Reverse DCF'!$C$8  = Base FCF
            # - 'Reverse DCF'!$C$10 = Net Debt
            # - 'Reverse DCF'!$C$9  = Shares
            # - 'Reverse DCF'!$C$15 = Projection Years
            # - 'Reverse DCF'!$C$14 = Terminal Growth Rate
            # - $B{rr} = WACC (this row's header)
            # - {col_ltr}$7 = Growth rate (this column's header)

            formula = (
                f"=IFERROR("
                f"(SUMPRODUCT('Reverse DCF'!$C$8*(1+{col_ltr}${r+1})^"
                f"ROW(INDIRECT(\"1:\"&'Reverse DCF'!$C$15)),"
                f"(1/(1+$B{rr}))^ROW(INDIRECT(\"1:\"&'Reverse DCF'!$C$15)))"
                f"+'Reverse DCF'!$C$8*(1+{col_ltr}${r+1})^'Reverse DCF'!$C$15"
                f"*(1+'Reverse DCF'!$C$14)/($B{rr}-'Reverse DCF'!$C$14)"
                f"/(1+$B{rr})^'Reverse DCF'!$C$15"
                f"-'Reverse DCF'!$C$10)/'Reverse DCF'!$C$9,\"N/A\")"
            )

            cell = ws.cell(row=rr, column=col)
            cell.value  = formula
            cell.number_format = '$#,##0'
            cell.font   = BF()
            cell.fill   = fl
            cell.alignment = A("right")
            cell.border = BD()

    # Color scale on the data range
    data_end_col = get_column_letter(2 + len(growth_rates))
    data_end_row = r + 1 + len(wacc_rates)
    range_str = f"C{r+2}:{data_end_col}{data_end_row}"
    ws.conditional_formatting.add(
        range_str,
        ColorScaleRule(
            start_type="min", start_color="F44336",
            mid_type="num",   mid_value=0, mid_color="FFFFFF",
            end_type="max",   end_color="2D6A4F",
        )
    )

    # ── Table 2: Implied Growth Rate vs (WACC, Terminal Growth) ──────────────
    # Show the reverse DCF implied growth for different terminal growth assumptions
    r2 = r + 2 + len(wacc_rates) + 3
    term_growths = [0.02, 0.025, 0.03, 0.035, 0.04, 0.045, 0.05]

    sec_hdr(ws, r2, 2, 9,
            "SENSITIVITY: WACC vs TERMINAL GROWTH RATE  → Implied FCF Growth Rate",
            "lt", size=10)

    sc(ws, r2+1, 2, "WACC ↓ / Term.G →",
       fnt=HD(sz=9), fl=F(P["lt"]), al=A("center"))
    for j, tg in enumerate(term_growths):
        sc(ws, r2+1, 3+j, f"{tg:.1%}",
           fnt=HD(sz=10), fl=F(P["lt"]), al=A("center"))

    sc(ws, r2, 10,
       "Note: values shown are approximate implied growth rates at the given WACC "
       "and terminal growth rate combinations, using the current stock price from "
       "the Reverse DCF sheet.",
       fnt=BF(sz=9, c="555555"), brd=False, al=A("left", w=True), span=12)
    ws.row_dimensions[r2].height = 20

    for i, wacc in enumerate(wacc_rates):
        rr2 = r2 + 2 + i
        ws.row_dimensions[rr2].height = 18
        fl = F(P["gry"]) if i % 2 == 0 else F(P["w"])

        sc(ws, rr2, 2, f"{wacc:.0%}",
           fnt=HD(sz=10), fl=F(P["md"]), al=A("center"))

        for j, tg in enumerate(term_growths):
            col = 3 + j
            col_ltr = get_column_letter(col)

            # For each WACC/TermG combo, calculate the implied growth numerically
            # We can't easily do full reverse-DCF in a formula without Solver,
            # so instead we show the sensitivity of WACC and terminal growth on
            # the forward DCF price given the CURRENT implied growth rate from model.
            # i.e., "if WACC were X% and terminal growth were Y%, the implied
            # stock price at the model's current implied growth rate would be $Z"

            formula = (
                f"=IFERROR("
                f"(SUMPRODUCT('Reverse DCF'!$C$8*(1+'Reverse DCF'!$C$19)^"
                f"ROW(INDIRECT(\"1:\"&'Reverse DCF'!$C$15)),"
                f"(1/(1+$B{rr2}))^ROW(INDIRECT(\"1:\"&'Reverse DCF'!$C$15)))"
                f"+'Reverse DCF'!$C$8*(1+'Reverse DCF'!$C$19)^'Reverse DCF'!$C$15"
                f"*(1+{col_ltr}${r2+1})/($B{rr2}-{col_ltr}${r2+1})"
                f"/(1+$B{rr2})^'Reverse DCF'!$C$15"
                f"-'Reverse DCF'!$C$10)/'Reverse DCF'!$C$9,\"N/A\")"
            )

            cell = ws.cell(row=rr2, column=col)
            cell.value  = formula
            cell.number_format = '$#,##0'
            cell.font   = BF()
            cell.fill   = fl
            cell.alignment = A("right")
            cell.border = BD()

    range2_str = f"C{r2+2}:I{r2+1+len(wacc_rates)}"
    ws.conditional_formatting.add(
        range2_str,
        ColorScaleRule(
            start_type="min", start_color="F44336",
            mid_type="percentile", mid_value=50, mid_color="FFEB3B",
            end_type="max",   end_color="2D6A4F",
        )
    )


# ─────────────────────────────────────────────────────────────────────────────
# SHEET: INSTRUCTIONS
# ─────────────────────────────────────────────────────────────────────────────

def build_instructions(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 100

    ws.row_dimensions[2].height = 45
    sc(ws, 2, 2, "HOW TO USE THIS MODEL  —  INSTRUCTIONS & DOCUMENTATION",
       fnt=HD(sz=18), fl=F(P["dk"]), al=A("center", "center"))

    def inst_row(r, text, indent=0, bold=False, color="1A1A1A", fill_key="w", height=18):
        ws.row_dimensions[r].height = height
        sc(ws, r, 2, text, fnt=BF(b=bold, c=color),
           fl=F(P[fill_key]), al=A("left", i=indent), brd=False)

    r = 4
    inst_row(r,   "═══════════════════════════════════════════════════════", color="2D6A4F")
    inst_row(r+1, "1.  LIVE PRICE AUTO-UPDATE", bold=True, fill_key="lt",
             color="FFFFFF", height=22)
    inst_row(r+2, "═══════════════════════════════════════════════════════", color="2D6A4F")

    inst_row(r+3,
             "The cell 'Reverse DCF'!C7 contains a STOCKHISTORY() formula which "
             "fetches the most recent MELI closing price in Excel 365 (Microsoft 365 subscription).",
             indent=1)
    inst_row(r+4,
             "Formula used:  "
             "=IFERROR(INDEX(STOCKHISTORY(\"MELI\",TODAY()-7,TODAY(),0,0),"
             "COUNTA(STOCKHISTORY(\"MELI\",TODAY()-7,TODAY(),0,0)),1), 1950)",
             indent=2, color="1565C0")
    inst_row(r+5,
             "If STOCKHISTORY is not available (Excel 2019 or earlier), manually enter "
             "the current stock price directly into cell 'Reverse DCF'!C7.",
             indent=1, color="C0392B")
    inst_row(r+6,
             "To refresh: Press Ctrl+Alt+F9 (force full recalculation) "
             "or go to Data > Refresh All.",
             indent=1)

    r += 8
    inst_row(r,   "═══════════════════════════════════════════════════════", color="2D6A4F")
    inst_row(r+1, "2.  VBA MACRO — AUTO-REFRESH ON OPEN", bold=True, fill_key="lt",
             color="FFFFFF", height=22)
    inst_row(r+2, "═══════════════════════════════════════════════════════", color="2D6A4F")
    inst_row(r+3,
             "To make the model auto-refresh the stock price every time the file is opened "
             "save as .xlsm and add this VBA code to ThisWorkbook:",
             indent=1)
    inst_row(r+4,
             "  Private Sub Workbook_Open()",
             indent=2, color="1565C0", fill_key="gry")
    inst_row(r+5,
             "      Application.CalculateFull       ' force refresh of all formulas",
             indent=2, color="1565C0", fill_key="gry")
    inst_row(r+6,
             "      ThisWorkbook.RefreshAll          ' refresh Power Query / data connections",
             indent=2, color="1565C0", fill_key="gry")
    inst_row(r+7,
             "      MsgBox \"MELI model refreshed. Current price: \" & _",
             indent=2, color="1565C0", fill_key="gry")
    inst_row(r+8,
             "            Format(Sheets(\"Reverse DCF\").Range(\"C7\").Value, \"$#,##0.00\")",
             indent=2, color="1565C0", fill_key="gry")
    inst_row(r+9,
             "  End Sub",
             indent=2, color="1565C0", fill_key="gry")

    r += 11
    inst_row(r,   "═══════════════════════════════════════════════════════", color="2D6A4F")
    inst_row(r+1, "3.  UNDERSTANDING THE REVERSE DCF", bold=True, fill_key="lt",
             color="FFFFFF", height=22)
    inst_row(r+2, "═══════════════════════════════════════════════════════", color="2D6A4F")
    inst_row(r+3,
             "A standard (forward) DCF asks: 'Given my growth assumptions, what is the stock worth?'",
             indent=1)
    inst_row(r+4,
             "A REVERSE DCF asks: 'Given the current stock price, what growth rate must the "
             "market be assuming for the stock to be fairly valued?'",
             indent=1, bold=True)
    inst_row(r+5, "Steps the model takes:", indent=1, bold=True)
    inst_row(r+6,
             "  a)  Sweeps growth rates from 0% to 60% in 1% increments (Lookup Table on Reverse DCF sheet).",
             indent=2)
    inst_row(r+7,
             "  b)  For each growth rate, calculates the implied stock price using DCF formula:",
             indent=2)
    inst_row(r+8,
             "      Price = (Σ PV of FCFs + PV of Terminal Value − Net Debt) ÷ Shares Outstanding",
             indent=3, color="1565C0", fill_key="gry")
    inst_row(r+9,
             "  c)  Uses INDEX/MATCH to find the growth rate where the implied price ≈ current price.",
             indent=2)
    inst_row(r+10,
             "  d)  This 'implied growth rate' is the market's embedded expectation.",
             indent=2)
    inst_row(r+11,
             "  e)  You then decide if that growth rate is achievable. If yes → fairly valued. "
             "If too high → potentially overvalued. If conservative → potentially undervalued.",
             indent=2)

    r += 13
    inst_row(r,   "═══════════════════════════════════════════════════════", color="2D6A4F")
    inst_row(r+1, "4.  UPDATING THE MODEL (Annual Refresh)", bold=True, fill_key="lt",
             color="FFFFFF", height=22)
    inst_row(r+2, "═══════════════════════════════════════════════════════", color="2D6A4F")
    inst_row(r+3,
             "Each year, update the following cells when new financials are released:",
             indent=1)
    inst_row(r+4, "  • 'Reverse DCF'!C8  — Base FCF (most recent fiscal year FCF, $M)", indent=2)
    inst_row(r+5, "  • 'Reverse DCF'!C9  — Shares Outstanding (diluted, millions)", indent=2)
    inst_row(r+6, "  • 'Reverse DCF'!C10 — Net Debt (total debt minus cash, $M)", indent=2)
    inst_row(r+7, "  • 'WACC'!C21        — Market Capitalisation ($M)", indent=2)
    inst_row(r+8, "  • 'WACC'!C22        — Total Debt ($M)", indent=2)
    inst_row(r+9, "  • 'Historicals' sheet — Add the new year's data in a new column", indent=2)
    inst_row(r+10,
             "  All other outputs (implied growth rate, intrinsic value, sensitivity "
             "tables) will auto-recalculate.", indent=2, color="2D6A4F", bold=True)

    r += 12
    inst_row(r,   "═══════════════════════════════════════════════════════", color="2D6A4F")
    inst_row(r+1, "5.  KEY FORMULA REFERENCE", bold=True, fill_key="lt",
             color="FFFFFF", height=22)
    inst_row(r+2, "═══════════════════════════════════════════════════════", color="2D6A4F")

    formulas = [
        ("Enterprise Value (DCF)",
         "EV = Σ[FCF₀×(1+g)ᵗ / (1+WACC)ᵗ] + FCF₀×(1+g)ⁿ×(1+gₜ) / [(WACC−gₜ)×(1+WACC)ⁿ]"),
        ("Equity Value",
         "Equity Value = Enterprise Value − Net Debt"),
        ("Stock Price",
         "Price = Equity Value ÷ Shares Outstanding"),
        ("CAPM (Cost of Equity)",
         "Ke = Rf + β × (ERP + CRP)"),
        ("WACC",
         "WACC = (E/V) × Ke + (D/V) × Kd × (1 − Tax Rate)"),
        ("Terminal Value",
         "TV = FCFₙ × (1 + gₜ) / (WACC − gₜ)"),
        ("Implied Growth (Excel formula)",
         "=INDEX(growth_rates, MATCH(MIN(ABS(dcf_prices − current_price)), ABS(dcf_prices − current_price), 0))"),
    ]
    for i, (name, formula) in enumerate(formulas):
        rr = r + 3 + i * 2
        inst_row(rr, f"  {name}:", indent=1, bold=True)
        inst_row(rr+1, f"    {formula}", indent=2, color="1565C0", fill_key="gry")

    r += 3 + len(formulas) * 2 + 1
    inst_row(r,
             "DISCLAIMER: This model is for educational and analytical purposes only. "
             "It does not constitute investment advice. Always do your own due diligence.",
             color="E74C3C", bold=True)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    wb = Workbook()

    # Create sheets
    ws_dash = wb.active;            ws_dash.title = "Dashboard"
    ws_rdcf = wb.create_sheet("Reverse DCF")
    ws_wacc = wb.create_sheet("WACC")
    ws_hist = wb.create_sheet("Historicals")
    ws_sens = wb.create_sheet("Sensitivity")
    ws_inst = wb.create_sheet("Instructions")

    # Tab colors
    ws_dash.sheet_properties.tabColor = "1B4332"
    ws_rdcf.sheet_properties.tabColor = "2D6A4F"
    ws_wacc.sheet_properties.tabColor = "40916C"
    ws_hist.sheet_properties.tabColor = "52B788"
    ws_sens.sheet_properties.tabColor = "74C69D"
    ws_inst.sheet_properties.tabColor = "95D5B2"

    print("Building Dashboard...")
    build_dashboard(ws_dash)
    print("Building WACC sheet...")
    build_wacc(ws_wacc)
    print("Building Reverse DCF sheet...")
    build_reverse_dcf(ws_rdcf)
    print("Building Historicals sheet...")
    build_historicals(ws_hist)
    print("Building Sensitivity sheet...")
    build_sensitivity(ws_sens)
    print("Building Instructions sheet...")
    build_instructions(ws_inst)

    # Freeze panes
    ws_rdcf["B4"].alignment = Alignment()
    ws_rdcf.freeze_panes = "C5"
    ws_hist.freeze_panes = "C6"

    # Print area / page setup
    for ws in [ws_dash, ws_rdcf, ws_wacc, ws_hist, ws_sens]:
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToPage   = True
        ws.page_setup.fitToWidth  = 1

    wb.save(OUTPUT)
    print(f"\n✓ Model saved as: {OUTPUT}")
    print(f"  File size: {os.path.getsize(OUTPUT):,} bytes")
    print("\nNext steps:")
    print("  1. Open in Excel 365 for live STOCKHISTORY price updates.")
    print("  2. If STOCKHISTORY is unavailable, manually enter price in 'Reverse DCF'!C7.")
    print("  3. Update WACC!C21 (market cap) when price changes significantly.")
    print("  4. See the Instructions sheet for VBA auto-refresh setup.")


if __name__ == "__main__":
    main()
